# -*- coding: utf-8 -*-

import csv
import io
import sys
from openpyxl import Workbook
from pyloco import Task

class Pydict2xlsx(Task):
    """converts Python dictionary to Microsoft Excel file

'pydict2xlsx' task converts a Python dictionary to Microsoft Excel file
or CSV file.

Example(s)
----------

Current version of the task assumes that 'docx2text' pyloco task is used as
a previous task as shown below.

First, make sure that 'docx2text' task is installed by running the following
command. ::

    >>> pyloco docx2text -h

You will see a help message of 'docx2text' on screen. If failed, please use
following command to install. ::

    >>> pyloco install docx2text

Follwoing command reads my.docx MS World file and convert tables in the file
to worksheets of MS Excel file. ::

    >>> pyloco docx2text my.docx -- pydict2xlsx -t xlsx
    tables.xlsx

Follwoing command reads my.docx MS World file and convert tables in the file
to CSV format text file. ::

    >>> pyloco docx2text my.docx -- pydict2xlsx -t csv
    tables.csv
"""

    _name_ = "pydict2xlsx"
    _version_ = "0.1.4"

    def __init__(self, parent):

        self.add_data_argument("data", type=str, help="input Python dictionary")

        self.add_option_argument("-t", "--type", metavar="type",
                default="xlsx", 
                help="output file format (default='xlsx')") 

        self.add_option_argument(
            "-o", "--output", help=("output file")
        )

        self.register_forward("data", help="output data")

    def perform(self, targs):

        tables = targs.data["table"]

        if targs.type == "xlsx":
            wb = Workbook()

            for tid, table in tables.items():
                if tid == 0:
                    ws = wb.active
                else:
                    ws = wb.create_sheet(str(tid))
                for rid, row in table.items():
                    for cid, cell in row.items():
                        c = ws.cell(row=rid+1, column=cid+1, value=cell)

            # Save the file
            outfile = targs.output if targs.output else "tables.xlsx"
            wb.save(outfile)

            self.add_forward(data=wb)

        elif targs.type == "csv":

            outfile = targs.output if targs.output else "tables.csv"
            with io.open(outfile, 'w', encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile, delimiter=',',
                                        quotechar='|', quoting=csv.QUOTE_MINIMAL)

                writer.writerow(["tableNum", "rowNum", "colNum", "Value"])
                for tid, table in tables.items():
                    for rid, row in table.items():
                        for cid, cell in row.items():
                            value = cell.replace("\n", " ").replace(",", " ")
                            writer.writerow([tid, rid, cid, value])

                self.add_forward(data=tables)
        else:
            print("Unknown output type: %s" % targs.type)
            sys.exit(1)
